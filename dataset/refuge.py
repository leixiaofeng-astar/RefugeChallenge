from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import logging
import os
from collections import OrderedDict

import numpy as np
from scipy.io import loadmat, savemat
from openpyxl import load_workbook
import csv
import cv2
import numpy as np

from dataset.dataset import FoveaDataset


logger = logging.getLogger(__name__)


class Dataset(FoveaDataset):
    def __init__(self, cfg, root, image_set, is_train, transform=None):
        super(Dataset, self).__init__(cfg, root, image_set, is_train, transform)
        self.cfg = cfg
        self.image_set = image_set
        self.db = self._get_db(image_set)

        # test_img = cv2.imread(self.db[0]['image'], cv2.IMREAD_COLOR)
        # test_img = self.db[0]['image']
        # self.db_image_size = np.array([test_img.shape[1], test_img.shape[0]])

        # import pdb
        # pdb.set_trace()
        self.db_image_size = []
        db_len = len(self.db)
        for i in range(db_len):
            test_img = self.db[i]['image']
            self.db_image_size.append(np.array([test_img.shape[1], test_img.shape[0]]))

        if is_train and cfg.DATASET.TRAIN_FOLD > 0:
            np.random.seed(1234) # fix seed
            perm = np.random.permutation(len(self.db))
            perm = np.split(perm, 5, axis=0)
            del perm[cfg.DATASET.TRAIN_FOLD - 1] # remove this fold
            indices = np.concatenate(perm, axis=0)
            db = [self.db[_] for _ in indices]
            self.db = db

        logger.info('=> load {} samples'.format(len(self.db)))

    def is_image_file(self, img_path):
        img_name = os.path.basename(img_path)
        _, ext = os.path.splitext(img_name)
        if ext == '.jpg':
            return True
        else:
            return False

    def _get_db(self, image_set):
        TRAIN_AGE_MODEL = True
        if image_set == 'train':
            # training images and labels
            if TRAIN_AGE_MODEL:
                train_anno_filename = os.path.join(self.root, 'AGE-train-GT', 'Train_Fovea_locations_AGE.csv')
            else:
                train_anno_filename = os.path.join(self.root, 'Annotation-Training400',
                                                   'Annotation-Training400', 'Fovea_location.xlsx')
            workbook = load_workbook(train_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            train_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=3).value) - 1
                fy = float(booksheet.cell(row=i, column=4).value) - 1
                fname = booksheet.cell(row=i, column=2).value
                if TRAIN_AGE_MODEL:
                    image_file = os.path.join(self.root, 'AGE-train', fname)
                else:
                    if fname[0] == 'n':
                        image_file = os.path.join(self.root, 'REFUGE-Training400', 'Training400', 'Non-Glaucoma', fname)
                    elif fname[0] == 'g':
                        image_file = os.path.join(self.root, 'REFUGE-Training400', 'Training400', 'Glaucoma', fname)
                    else:
                        assert False, 'unkown entry: %s' % (fname)
                if not self.is_image_file(image_file): continue

                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                train_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })
            return train_db
        elif image_set == 'test':
            # test images and labels
            # TODO
            if TRAIN_AGE_MODEL:
                test_anno_filename = os.path.join(self.root, 'AGE-test-GT', 'test_Fovea_locations_AGE.csv')
            else:
                make_sample_gt = False
                if make_sample_gt:
                    test_anno_filename = os.path.join(self.root, 'test-GT', 'test_Fovea_locations.xlsx')
                else:
                    # test_anno_filename = os.path.join(self.root, 'REFUGE-Test-GT', 'Glaucoma_label_and_Fovea_location.xlsx')
                    test_anno_filename = os.path.join(self.root, 'REFUGE-Validation400-GT', 'Fovea_locations_Val2Val.xlsx')
            workbook = load_workbook(test_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            test_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=4).value) - 1
                fy = float(booksheet.cell(row=i, column=5).value) - 1
                fname = booksheet.cell(row=i, column=2).value

                if TRAIN_AGE_MODEL:
                    image_file = os.path.join(self.root, 'AGE-test', fname)
                else:
                    if make_sample_gt:
                        image_file = os.path.join(self.root, 'test_img', fname)
                    else:
                        # image_file = os.path.join(self.root, 'REFUGE-Test400', 'Test400', fname)
                        image_file = os.path.join(self.root, 'REFUGE-Validation400', 'REFUGE-Val2Val', fname)
                    if not self.is_image_file(image_file): continue

                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                test_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })
            return test_db
        elif image_set == 'val':
            # validation images and labels
            # TODO
            if TRAIN_AGE_MODEL:
                val_anno_filename = os.path.join(self.root, 'AGE-val-GT', 'val_Fovea_locations_AGE.csv')
            else:
                val_anno_filename = os.path.join(self.root, 'REFUGE-Validation400-GT', 'Fovea_locations_Val2Train.xlsx')

            workbook = load_workbook(val_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            val_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=4).value) - 1
                fy = float(booksheet.cell(row=i, column=5).value) - 1
                fname = booksheet.cell(row=i, column=2).value
                if TRAIN_AGE_MODEL:
                    image_file = os.path.join(self.root, 'AGE-val', fname)
                else:
                    image_file = os.path.join(self.root, 'REFUGE-Validation400', 'REFUGE-Val2Train', fname)
                if not self.is_image_file(image_file): continue

                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                val_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })
            return val_db
        elif image_set == 'train+val':
            # training images and labels
            train_anno_filename = os.path.join(self.root, 'Annotation-Training400',
                                               'Annotation-Training400', 'Fovea_location.xlsx')
            workbook = load_workbook(train_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            train_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=3).value) - 1
                fy = float(booksheet.cell(row=i, column=4).value) - 1
                fname = booksheet.cell(row=i, column=2).value
                if fname[0] == 'n':
                    image_file = os.path.join(self.root, 'REFUGE-Training400', 'Training400', 'Non-Glaucoma', fname)
                elif fname[0] == 'g':
                    image_file = os.path.join(self.root, 'REFUGE-Training400', 'Training400', 'Glaucoma', fname)
                else:
                    assert False, 'unkown entry: %s' % (fname)
                if not self.is_image_file(image_file): continue

                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                train_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })

            # validation images and labels
            val_anno_filename = os.path.join(self.root, 'REFUGE-Validation400-GT', 'Fovea_locations.xlsx')
            workbook = load_workbook(val_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            val_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=4).value) - 1
                fy = float(booksheet.cell(row=i, column=5).value) - 1
                fname = booksheet.cell(row=i, column=2).value
                image_file = os.path.join(self.root, 'REFUGE-Validation400', 'REFUGE-Validation400', fname)
                if not self.is_image_file(image_file): continue

                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                val_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })
            return train_db + val_db
        elif image_set == 'val2':
            # TODO: refuge2 validation db -- xiaofeng comment
            # refuge2 validation images and labels
            trial_enable = self.cfg.TEST.TRIAL_RUN
            if not trial_enable:
                val2_anno_filename = os.path.join(self.root, 'REFUGE-Test-GT', 'Glaucoma_label_and_Fovea_location.xlsx')
                # val2_anno_filename = os.path.join(self.root, 'Refuge2-Validation-GT', 'Fovea_locations_dummy.xlsx')
            else:
                # prepare for refuge2 final submission
                val2_anno_filename = os.path.join(self.root, 'Refuge2-Ext-GT', 'IDRiD_Fovea_Center_GT.xlsx')

            workbook = load_workbook(val2_anno_filename)
            booksheet = workbook.active
            rows = booksheet.rows
            columns = booksheet.columns
            val2_db = []
            for i, row in enumerate(rows, 1):
                if i == 1: continue  # skip the first row
                # substract 1 pixel as we assume indexing from zero
                fx = float(booksheet.cell(row=i, column=4).value) - 1
                fy = float(booksheet.cell(row=i, column=5).value) - 1
                fname = booksheet.cell(row=i, column=2).value
                if not trial_enable:
                    image_file = os.path.join(self.root, 'REFUGE-Test400', 'Test400', fname)
                    # image_file = os.path.join(self.root, 'Refuge2-Validation', 'Refuge2-Validation', fname)
                else:
                    # prepare for refuge2 final submission
                    fname = fname + '.jpg'
                    image_file = os.path.join(self.root, 'Refuge2-Ext', 'Refuge2-Ext', fname)

                if not self.is_image_file(image_file): continue
                data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                val2_db.append({
                    'image': data_numpy,
                    'fovea': np.array([fx, fy], np.float32),
                    'filename': image_file,
                })
            return val2_db
        else:
            assert ('Unknown image set: %s' % (image_set))


    def evaluate(self, preds, output_dir, debug_enable=False):
        num_images = len(self.db)
        assert num_images == len(self.db)

        # the predicted coordinates are based on the resized and center-cropped
        # images, convert them back
        image_size = self.cfg.MODEL.IMAGE_SIZE
        crop_size = self.cfg.MODEL.CROP_SIZE
        pw = (image_size[0] - crop_size[0]) // 2
        ph = (image_size[1] - crop_size[1]) // 2
        preds[:, 0] += pw
        preds[:, 1] += ph
        # preds[:, 0] *= (self.db_image_size[0] * 1.0 / image_size[0])
        # preds[:, 1] *= (self.db_image_size[1] * 1.0 / image_size[1])

        for x in range(num_images):
            preds[x, 0] *= (self.db_image_size[x][0] * 1.0 / image_size[0])
            preds[x, 1] *= (self.db_image_size[x][1] * 1.0 / image_size[1])
            # self.db_image_size.append(np.array([test_img.shape[1], test_img.shape[0]]))

        l2_dist_sum = 0.
        for _ in range(num_images):
            gt = self.db[_]['fovea']
            l2_dist_sum += np.sqrt(np.sum((preds[_, :] - gt)**2))
        l2_dist_avg = l2_dist_sum / num_images

        if output_dir is not None:
            csv_file = os.path.join(output_dir, 'fovea_location_results.csv')
            with open(csv_file, 'w') as f:
                cw = csv.writer(f, delimiter=",", lineterminator="\n")
                if debug_enable:
                    cw.writerow(['ImageName', 'Fovea_X', 'Fovea_Y', 'FoveaX_GT', 'FoveaY_GT', 'X_bias', 'Y_bias'])
                    for _ in range(num_images):
                        image_name = os.path.basename(self.db[_]['filename'])
                        fovea_x = '%.2f' % (preds[_, 0]+1)
                        fovea_y = '%.2f' % (preds[_, 1]+1)
                        FoveaX_GT = '%.2f' % (self.db[_]['fovea'][0])
                        FoveaY_GT = '%.2f' % (self.db[_]['fovea'][1])
                        X_bias = '%.2f' % (preds[_, 0] - self.db[_]['fovea'][0])
                        Y_bias = '%.2f' % (preds[_, 1] - self.db[_]['fovea'][1])
                        cw.writerow([image_name, fovea_x, fovea_y, FoveaX_GT, FoveaY_GT, X_bias, Y_bias])
                else:
                    cw.writerow(['ImageName', 'Fovea_X', 'Fovea_Y'])
                    for _ in range(num_images):
                        image_name = os.path.basename(self.db[_]['filename'])
                        # TODO : check if +1 if necessary
                        fovea_x = '%.2f' %(preds[_, 0] + 1)
                        fovea_y = '%.2f' %(preds[_, 1] + 1)
                        cw.writerow([image_name, fovea_x, fovea_y])
                f.close()

        return l2_dist_avg
